from openpyxl import load_workbook
from openpyxl import Workbook
# import numpy as np
import pandas as pd

# input file 입력

wb2 = Workbook()
ws2 = wb2.active
filename2 = 'input.xlsx'
ws2.title = 'input'

ws2['A1'] = '학생 이름'
ws2['B1'] = '학생 답'
ws2['C1'] = '문제 답'
ws2['D1'] = '성적'
ws2['E1'] = '등수'
ws2['F1'] = '30% flag'

std_n = int(input("학생 수를 입력하세요: "))
ans = input("정답을 순서대로 입력하세요: ")
scr = [0 for i in range(0, std_n)]
std_name = [0 for i in range(0, std_n)]

for i in range(1, std_n + 1):
    std_name[i-1] = input(str(i) + "번째 학생 이름을 입력하세요: ")
    ws2.cell(row=i + 1, column=1).value = std_name[i-1]
    std_ans = input(str(i) + "번째 학생 답을 순서대로 입력하세요: ")
    ws2.cell(row=i + 1, column=2, value=std_ans)
    ws2.cell(row=i + 1, column=3, value=ans)
    for j in range(0, len(std_ans)):
        if std_ans[j] == ans[j]:
            scr[i - 1] += 1
    ws2.cell(row=i + 1, column=4).value = scr[i - 1] / len(std_ans) * 100

df = pd.DataFrame({'number': [0 for i in range(0, std_n)], 'score': scr}, columns=['number', 'score'])
df['rank'] = df['score'].rank(method='min', ascending=False)
# print(df['rank'][0])

flag = [0 for i in range(0, std_n)]
for i in range(1, std_n + 1):
    ws2.cell(row=i + 1, column=5).value = df['rank'][i - 1]
    if df['rank'][i - 1] < std_n * 0.3:
        flag[i - 1] = 1
    ws2.cell(row=i + 1, column=6).value = flag[i - 1]
# for i in range(1,std_n+1):

wb2.save(filename=filename2)

# 완료

# ouput file 기본 정보 입력
wb = Workbook()
ws = wb.active
filename = 'output.xlsx'
ws.title = 'output'

ws['A1'] = '문제 번호'
n = int(input("문제 개수를 입력하세요: "))
for i in range(1, n + 1):
    ws.cell(row=i + 1, column=1).value = i

ws['B1'] = 1
ws['C1'] = 2
ws['D1'] = 3
ws['E1'] = 4
ws['F1'] = 5
ws['G1'] = '전체 정답률'
ws['H1'] = '상위 30% 정답률'

# 읽어들이기 + main 계산
load_wb = load_workbook(filename='input.xlsx')
load_ws = load_wb['input']

# 각 보기별 선택률 (ch1~5) # 정답률 ch6
# std_n = load_ws.cell(row=2, column=2).value
# std_n=10
ch1 = 0
ch2 = 0
ch3 = 0
ch4 = 0
ch5 = 0
ch6 = 0
ch7 = 0
# std_ans_len = len(list(load_ws.cell(row=2, column=2).value))
std_ans_len = n
# ans = list(load_ws.cell(row=2, column=3).value)

p = [[0 for i in range(0, std_ans_len)] for j in range(0, std_n)]
scr_dist = [0 for i in range(0, 5)]

for i in range(1, std_n + 1):
    std_ans = list(load_ws.cell(row=i + 1, column=2).value)
    for j in range(0, std_ans_len):
        p[i - 1][j] = std_ans[j]
#     print("p[",i-1,"][",j,"] = ",p[i-1][j])


for i in range(1, std_n + 1):
    scr = load_ws.cell(row=i + 1, column=4).value
    # print(scr)
    if scr < 20:
        scr_dist[0] += 1
    elif 20 <= scr < 40:
        scr_dist[1] += 1
    elif 40 <= scr < 60:
        scr_dist[2] += 1
    elif 60 <= scr < 80:
        scr_dist[3] += 1
    elif scr >= 80:
        scr_dist[4] += 1

# print("scr_dist",scr_dist)
ws.cell(row=n + 3, column=1).value = "분포"
ws.cell(row=n + 4, column=1).value = "0-20"
ws.cell(row=n + 5, column=1).value = "20-40"
ws.cell(row=n + 6, column=1).value = "40-60"
ws.cell(row=n + 7, column=1).value = "60-80"
ws.cell(row=n + 8, column=1).value = "80-100"
ws.cell(row=n + 4, column=2).value = scr_dist[0] / std_n * 100
ws.cell(row=n + 5, column=2).value = scr_dist[1] / std_n * 100
ws.cell(row=n + 6, column=2).value = scr_dist[2] / std_n * 100
ws.cell(row=n + 7, column=2).value = scr_dist[3] / std_n * 100
ws.cell(row=n + 8, column=2).value = scr_dist[4] / std_n * 100

for k in range(0, std_ans_len):
    for i in range(0, std_n):
        j = p[i][k]
        if j == '1':
            ch1 += 1
        elif j == '2':
            ch2 += 1
        elif j == '3':
            ch3 += 1
        elif j == '4':
            ch4 += 1
        elif j == '5':
            ch5 += 1

        if j == ans[k]:
            ch6 += 1

    ws.cell(row=k + 2, column=2).value = ch1 / std_n * 100
    ws.cell(row=k + 2, column=3).value = ch2 / std_n * 100
    ws.cell(row=k + 2, column=4).value = ch3 / std_n * 100
    ws.cell(row=k + 2, column=5).value = ch4 / std_n * 100
    ws.cell(row=k + 2, column=6).value = ch5 / std_n * 100
    ws.cell(row=k + 2, column=7).value = ch6 / std_n * 100

    ch1 = 0
    ch2 = 0
    ch3 = 0
    ch4 = 0
    ch5 = 0
    ch6 = 0

# 상위 30프로 정답률-> 성적을 내서 30프로 넘는 학생에 flag
best_std_n = 0
f = [0 for i in range(0, std_n)]
for m in range(0, std_n):
    flag = load_ws.cell(row=m + 2, column=6).value
    if flag == 1:
        f[best_std_n] = m
        best_std_n += 1
# print(best_std_n)

for m in range(0, std_n):
    flag = load_ws.cell(row=m + 2, column=6).value
    if flag == 1:
        for k in range(0, std_ans_len):
            for i in range(0, best_std_n):
                j = p[f[i]][k]
                # print("j=",j)
                if j == ans[k]:
                    ch7 += 1
            # print("ch7=",ch7)
            ws.cell(row=k + 2, column=8).value = ch7 / best_std_n * 100
            ch7 = 0
    else:
        continue

# 정오표
for m in range(0, std_n):
    # ws.cell(row=1, column=m + 10).value = m + 1
    ws.cell(row=1, column=m + 10).value = std_name[m]
    std_ans = list(load_ws.cell(row=m + 2, column=2).value)
    for k in range(0, len(std_ans)):
        if std_ans[k] == ans[k]:
            ws.cell(row=k + 2, column=m + 10).value = 'O'
        else:
            ws.cell(row=k + 2, column=m + 10).value = 'X'

# 완료

wb.save(filename=filename)

# wb.save("C:/Users/user/Desktop/test1.xlsx")